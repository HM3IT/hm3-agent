import logging
import time
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path


from typing import TYPE_CHECKING

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from config import TZ, LOG_LEVEL

if TYPE_CHECKING:
    from schema import ReceiptData


logging.basicConfig(
    level=LOG_LEVEL,
    format="%(asctime)s | %(levelname)s | %(message)s",
)
logger = logging.getLogger("finance_bot")


class ExcelLedger:
    HEADERS = [
        "Transaction Date",
        "Transaction Time",
        "Merchant",
        "Category",
        "Item Description",
        "Amount",
        "Currency",
        "Payment Method",
        "Receipt Number",
        "Tax Amount",
        "Confidence",
        "Notes",
        "Source Key",
        "Telegram Chat ID",
        "Telegram Message ID",
        "Original Filename",
        "Processed At",
    ]

    def __init__(self, path: Path) -> None:
        self.path = path
        if not path.exists():
            self._create()

    def _save_with_retry(
        self,
        workbook,
        attempts: int = 3,
        delay_seconds: int = 5,
    ) -> None:
        for attempt in range(1, attempts + 1):
            try:
                workbook.save(self.path)
                return
            except PermissionError:
                if attempt == attempts:
                    raise

                logger.warning(
                    "Excel file is locked. Save attempt %s/%s failed. "
                    "Retrying in %s seconds.",
                    attempt,
                    attempts,
                    delay_seconds,
                )
                time.sleep(delay_seconds)

    def _create(self) -> None:
        workbook = Workbook()
        expenses = workbook.active
        expenses.title = "Expenses"
        expenses.append(self.HEADERS)
        self._style_header(expenses)
        expenses.freeze_panes = "A2"
        expenses.auto_filter.ref = f"A1:{get_column_letter(len(self.HEADERS))}1"
        workbook.create_sheet("Daily Summary")
        workbook.create_sheet("Category Summary")
        self._save_with_retry(workbook)

    @staticmethod
    def _style_header(sheet) -> None:
        fill = PatternFill("solid", fgColor="1F4E78")
        for cell in sheet[1]:
            cell.fill = fill
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")

    def contains_source(self, source_key: str) -> bool:
        workbook = load_workbook(self.path, read_only=True, data_only=True)
        try:
            sheet = workbook["Expenses"]
            source_column = self.HEADERS.index("Source Key") + 1
            for row in sheet.iter_rows(
                min_row=2,
                min_col=source_column,
                max_col=source_column,
                values_only=True,
            ):
                if row[0] == source_key:
                    return True
            return False
        finally:
            workbook.close()

    def append(
        self,
        receipt: "ReceiptData",
        *,
        source_key: str,
        chat_id: int,
        message_id: int,
        filename: str,
    ) -> bool:
        if self.contains_source(source_key):
            return False

        workbook = load_workbook(self.path)
        sheet = workbook["Expenses"]
        sheet.append(
            [
                date.fromisoformat(receipt.transaction_date),
                receipt.transaction_time,
                receipt.merchant,
                receipt.category,
                receipt.item_description,
                float(receipt.total_amount),
                receipt.currency,
                receipt.payment_method,
                receipt.receipt_number,
                receipt.tax_amount,
                receipt.confidence,
                receipt.notes,
                source_key,
                chat_id,
                message_id,
                filename,
                datetime.now(TZ).replace(tzinfo=None),
            ]
        )

        last_row = sheet.max_row
        sheet.cell(last_row, 1).number_format = "yyyy-mm-dd"
        sheet.cell(last_row, 6).number_format = "#,##0.00"
        sheet.cell(last_row, 10).number_format = "#,##0.00"
        sheet.cell(last_row, 11).number_format = "0%"
        sheet.cell(last_row, 17).number_format = "yyyy-mm-dd hh:mm:ss"

        widths = {
            "A": 16,
            "B": 14,
            "C": 24,
            "D": 20,
            "E": 45,
            "F": 14,
            "G": 10,
            "H": 18,
            "I": 20,
            "J": 14,
            "K": 12,
            "L": 38,
            "M": 34,
            "N": 18,
            "O": 20,
            "P": 28,
            "Q": 22,
        }
        for column, width in widths.items():
            sheet.column_dimensions[column].width = width

        self._rebuild_summaries(workbook)
        self._save_with_retry(workbook)
        return True

    def _expense_rows(self, workbook) -> list[dict]:
        sheet = workbook["Expenses"]
        headers = [cell.value for cell in sheet[1]]
        rows = []
        for values in sheet.iter_rows(min_row=2, values_only=True):
            if not values[0]:
                continue
            rows.append(dict(zip(headers, values)))
        return rows

    def _rebuild_summaries(self, workbook) -> None:
        rows = self._expense_rows(workbook)

        daily: dict[tuple[date, str], Decimal] = {}
        categories: dict[tuple[str, str], Decimal] = {}

        for row in rows:
            raw_date = row["Transaction Date"]
            if isinstance(raw_date, datetime):
                transaction_date = raw_date.date()
            elif isinstance(raw_date, date):
                transaction_date = raw_date
            else:
                transaction_date = date.fromisoformat(str(raw_date))

            currency = str(row["Currency"])
            amount = Decimal(str(row["Amount"]))
            category = str(row["Category"])
            daily[(transaction_date, currency)] = (
                daily.get((transaction_date, currency), Decimal("0")) + amount
            )
            categories[(category, currency)] = (
                categories.get((category, currency), Decimal("0")) + amount
            )

        for name in ("Daily Summary", "Category Summary"):
            old_sheet = workbook[name]
            workbook.remove(old_sheet)

        daily_sheet = workbook.create_sheet("Daily Summary")
        daily_sheet.append(["Date", "Currency", "Total"])
        for (day, currency), amount in sorted(daily.items()):
            daily_sheet.append([day, currency, float(amount)])
        self._format_summary(daily_sheet, date_column=True)

        category_sheet = workbook.create_sheet("Category Summary")
        category_sheet.append(["Category", "Currency", "Total"])
        for (category, currency), amount in sorted(categories.items()):
            category_sheet.append([category, currency, float(amount)])
        self._format_summary(category_sheet, date_column=False)

    def _format_summary(self, sheet, *, date_column: bool) -> None:
        self._style_header(sheet)
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = "A1:C1"
        sheet.column_dimensions["A"].width = 22
        sheet.column_dimensions["B"].width = 12
        sheet.column_dimensions["C"].width = 16
        for row in range(2, sheet.max_row + 1):
            if date_column:
                sheet.cell(row, 1).number_format = "yyyy-mm-dd"
            sheet.cell(row, 3).number_format = "#,##0.00"

    def load_expenses(self) -> list[dict]:
        workbook = load_workbook(self.path, read_only=True, data_only=True)
        try:
            return self._expense_rows(workbook)
        finally:
            workbook.close()
